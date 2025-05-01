# -*- coding: utf-8 -*-
"""
OpenPyXL エンジンを使用したExcelエクスポーターの実装。
ExcelExporterクラスから使用されます。
"""

import os
import logging
import pandas as pd
import shutil

logger = logging.getLogger(__name__)

def export_openpyxl(exporter, data, output_path, template=None, **kwargs):
    """
    OpenPyXLを使用してExcelを出力
    
    Parameters
    ----------
    exporter : ExcelExporter
        エクスポーターインスタンス
    data : Any
        エクスポートするデータ
    output_path : str
        出力ファイルパス
    template : Optional[str]
        テンプレートファイルのパス
    **kwargs : dict
        追加のパラメータ
        
    Returns
    -------
    str
        出力ファイルのパス
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import LineChart, Reference
    except ImportError:
        exporter.add_error("OpenPyXLがインストールされていません。pip install openpyxl を実行してください。")
        return None
    
    # テンプレートを使用する場合
    if template and os.path.exists(template):
        try:
            # テンプレートファイルを開く
            shutil.copy(template, output_path)
            workbook = openpyxl.load_workbook(output_path)
        except Exception as e:
            exporter.add_warning(f"テンプレートの読み込みに失敗しました: {str(e)}")
            workbook = openpyxl.Workbook()
    else:
        # 新規ワークブック作成
        workbook = openpyxl.Workbook()
        
        # デフォルトシートの名前変更
        if "Sheet" in workbook.sheetnames:
            workbook.active.title = "Data"
        elif "Sheet1" in workbook.sheetnames:
            workbook.active.title = "Data"
    
    # メタデータと本体データを取得
    metadata = {}
    if hasattr(data, 'metadata') and isinstance(data.metadata, dict):
        metadata = data.metadata
    
    df = data.data
    
    # スタイルの設定
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    title_font = Font(bold=True, size=14)
    title_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # シート順序の決定
    sheet_layout = exporter.options.get("sheet_layout", ["metadata", "summary", "data", "wind_data", "strategy_points"])
    
    # データシートの出力
    if "data" in sheet_layout and exporter.options.get("include_data", True):
        if "Data" not in workbook.sheetnames:
            worksheet = workbook.create_sheet("Data")
        else:
            worksheet = workbook["Data"]
        
        # ヘッダーの書き込み
        for col_num, column_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num, value=column_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = header_border
            cell.alignment = header_alignment
        
        # データの書き込み
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)
        
        # 自動フィルターの設定
        if exporter.options.get("auto_filter", True):
            worksheet.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
        
        # 枠の固定
        if exporter.options.get("freeze_panes", True):
            worksheet.freeze_panes = "A2"
        
        # 列幅の調整
        for col_num, column_name in enumerate(df.columns, 1):
            column_width = max(len(str(column_name)), 10)
            worksheet.column_dimensions[get_column_letter(col_num)].width = column_width
    
    # メタデータシートの出力
    if "metadata" in sheet_layout and exporter.options.get("include_metadata", True) and metadata:
        if "Metadata" not in workbook.sheetnames:
            worksheet = workbook.create_sheet("Metadata")
        else:
            worksheet = workbook["Metadata"]
        
        # ヘッダーの書き込み
        header_cells = ["Key", "Value"]
        for col_num, header in enumerate(header_cells, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = header_border
            cell.alignment = header_alignment
        
        # メタデータの書き込み
        row_num = 2
        for key, value in metadata.items():
            worksheet.cell(row=row_num, column=1, value=key)
            worksheet.cell(row=row_num, column=2, value=str(value))
            row_num += 1
        
        # 列幅の調整
        worksheet.column_dimensions["A"].width = 30
        worksheet.column_dimensions["B"].width = 50
    
    # サマリーシートの出力
    if "summary" in sheet_layout and exporter.options.get("include_summary", True):
        if "Summary" not in workbook.sheetnames:
            worksheet = workbook.create_sheet("Summary")
        else:
            worksheet = workbook["Summary"]
        
        # ヘッダーの書き込み
        header_cells = ["Item", "Value"]
        for col_num, header in enumerate(header_cells, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = header_border
            cell.alignment = header_alignment
        
        # サマリーデータの作成と書き込み
        summary_data = exporter._create_summary_data(df)
        row_num = 2
        for key, value in summary_data.items():
            worksheet.cell(row=row_num, column=1, value=key)
            worksheet.cell(row=row_num, column=2, value=value)
            row_num += 1
        
        # 列幅の調整
        worksheet.column_dimensions["A"].width = 30
        worksheet.column_dimensions["B"].width = 30
    
    # 風データシートの出力
    if "wind_data" in sheet_layout and 'wind_speed' in df.columns and 'wind_direction' in df.columns:
        if "Wind Data" not in workbook.sheetnames:
            worksheet = workbook.create_sheet("Wind Data")
        else:
            worksheet = workbook["Wind Data"]
        
        # 風データの抽出
        wind_columns = ['timestamp', 'wind_speed', 'wind_direction']
        wind_data = df[wind_columns].copy()
        
        # ヘッダーの書き込み
        for col_num, column_name in enumerate(wind_columns, 1):
            cell = worksheet.cell(row=1, column=col_num, value=column_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = header_border
            cell.alignment = header_alignment
        
        # データの書き込み
        for row_idx, row in enumerate(wind_data.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)
        
        # 自動フィルターの設定
        if exporter.options.get("auto_filter", True):
            worksheet.auto_filter.ref = f"A1:{get_column_letter(len(wind_columns))}{len(wind_data) + 1}"
        
        # 枠の固定
        if exporter.options.get("freeze_panes", True):
            worksheet.freeze_panes = "A2"
        
        # 列幅の調整
        for col_num, column_name in enumerate(wind_columns, 1):
            column_width = max(len(str(column_name)), 12)
            worksheet.column_dimensions[get_column_letter(col_num)].width = column_width
        
        # チャートの追加（風速の時系列）
        if exporter.options.get("include_charts", True):
            try:
                # チャートシートの作成
                if "Wind Chart" not in workbook.sheetnames:
                    chart_sheet = workbook.create_sheet("Wind Chart")
                else:
                    chart_sheet = workbook["Wind Chart"]
                
                # チャートの作成
                chart = LineChart()
                chart.title = "Wind Speed Over Time"
                chart.x_axis.title = "Time"
                chart.y_axis.title = "Wind Speed"
                
                # データ範囲
                data = Reference(worksheet, min_col=2, min_row=1, max_row=len(wind_data) + 1, max_col=2)
                categories = Reference(worksheet, min_col=1, min_row=2, max_row=len(wind_data) + 1)
                
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(categories)
                
                # チャートの追加
                chart_sheet.add_chart(chart, "A1")
            except Exception as e:
                exporter.add_warning(f"風チャートの作成に失敗しました: {str(e)}")
    
    # ファイルの保存
    workbook.save(output_path)
    
    return output_path
